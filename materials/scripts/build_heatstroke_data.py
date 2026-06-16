#!/usr/bin/env python3
"""Build a clean heatstroke dataset: daily heat-emergency transports (FDMA, gov)
joined with daily temperature (Open-Meteo archive) for diverse prefectures.
Output: heatstroke.csv  (one row per prefecture x day, summer months May-Sep).
"""
import io, json, time, urllib.request, urllib.parse
import openpyxl

# year code -> (url dir, label year)
YEARS = {
    "h30": ("item", 2018),
    "r1":  ("items", 2019),
    "r2":  ("items", 2020),
    "r3":  ("items", 2021),
    "r4":  ("items", 2022),
    "r5":  ("items", 2023),
    "r6":  ("items", 2024),
}

# prefecture code -> (en, ja, lat, lon of capital)  diverse climates incl. Kobe
PREFS = {
    1:  ("Hokkaido", "北海道", 43.0618, 141.3545),   # Sapporo
    13: ("Tokyo",    "東京都", 35.6895, 139.6917),
    23: ("Aichi",    "愛知県", 35.1815, 136.9066),   # Nagoya
    27: ("Osaka",    "大阪府", 34.6937, 135.5023),
    28: ("Hyogo",    "兵庫県", 34.6913, 135.1830),   # Kobe (course location)
    40: ("Fukuoka",  "福岡県", 33.5902, 130.4017),
    47: ("Okinawa",  "沖縄県", 26.2124, 127.6809),   # Naha
}

def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urllib.request.urlopen(req, timeout=60).read()

# ---- 1) transports from FDMA xlsx -------------------------------------------
# transports[(year, month, day, pref_code)] = {"total":, "elderly":, "severe":}
transports = {}
for code, (d, year) in YEARS.items():
    url = f"https://www.fdma.go.jp/disaster/heatstroke/{d}/heatstroke003_data_{code}.xlsx"
    print("downloading", url)
    raw = fetch(url)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        # find columns by name
        idx = {name: i for i, name in enumerate(header)}
        c_date = idx.get("日付")
        c_pref = idx.get("都道府県コード")
        c_tot = idx.get("搬送人員（計）")
        c_eld = idx.get("年齢区分：高齢者")
        c_sev = idx.get("傷病程度：重症")
        c_dea = idx.get("傷病程度：死亡")
        for row in rows:
            if c_date is None or row[c_date] is None:
                continue
            dt = row[c_date]
            pc = row[c_pref]
            if pc not in PREFS:
                continue
            try:
                pc = int(pc)
            except Exception:
                continue
            key = (dt.year, dt.month, dt.day, pc)
            tot = row[c_tot] or 0
            eld = row[c_eld] or 0 if c_eld is not None else 0
            sev = (row[c_sev] or 0) if c_sev is not None else 0
            dea = (row[c_dea] or 0) if c_dea is not None else 0
            transports[key] = {"total": int(tot), "elderly": int(eld),
                               "severe": int(sev) + int(dea)}
    wb.close()
print("transport rows:", len(transports))

# ---- 2) temperature from Open-Meteo archive ---------------------------------
# temp[(pref_code, "YYYY-MM-DD")] = (tmax, tmean, humidity)
temp = {}
for pc, (en, ja, lat, lon) in PREFS.items():
    q = urllib.parse.urlencode({
        "latitude": lat, "longitude": lon,
        "start_date": "2018-05-01", "end_date": "2024-09-30",
        "daily": "temperature_2m_max,temperature_2m_mean,relative_humidity_2m_mean",
        "timezone": "Asia/Tokyo",
    })
    url = "https://archive-api.open-meteo.com/v1/archive?" + q
    print("weather", en)
    data = json.loads(fetch(url))
    d = data["daily"]
    for t, tmax, tmean, rh in zip(d["time"], d["temperature_2m_max"],
                                  d["temperature_2m_mean"], d["relative_humidity_2m_mean"]):
        temp[(pc, t)] = (tmax, tmean, rh)
    time.sleep(1)
print("temp rows:", len(temp))

# ---- 3) join + write --------------------------------------------------------
import csv
WD = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
import datetime
out = []
for (y, m, dd, pc), v in transports.items():
    iso = f"{y:04d}-{m:02d}-{dd:02d}"
    tw = temp.get((pc, iso))
    if tw is None:
        continue
    tmax, tmean, rh = tw
    if tmax is None:
        continue
    en, ja, *_ = PREFS[pc]
    wd = WD[datetime.date(y, m, dd).weekday()]
    out.append({
        "date": iso, "year": y, "month": m, "weekday": wd,
        "pref_code": pc, "pref_en": en, "pref_ja": ja,
        "tmax_c": round(tmax, 1), "tmean_c": round(tmean, 1),
        "humidity_pct": round(rh) if rh is not None else "",
        "transported": v["total"], "elderly": v["elderly"], "severe": v["severe"],
    })
out.sort(key=lambda r: (r["date"], r["pref_code"]))
cols = ["date", "year", "month", "weekday", "pref_code", "pref_en", "pref_ja",
        "tmax_c", "tmean_c", "humidity_pct", "transported", "elderly", "severe"]
with open("/tmp/heatstroke.csv", "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=cols)
    w.writeheader()
    w.writerows(out)
print("wrote", len(out), "rows -> /tmp/heatstroke.csv")
